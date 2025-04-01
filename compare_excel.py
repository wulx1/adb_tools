import datetime
import os
from openpyxl import load_workbook
from tkinter import filedialog, messagebox, Toplevel
from tkinter import *
import tkinter as tk


def save_to_txt_file(file_path, differences):
    with open(file_path, 'w', encoding='utf-8') as f:
        f.write("比较结果：\n")
        f.write("\n".join(differences))


def compare_excel_files(file1, file2):
    wb1 = load_workbook(file1)
    wb2 = load_workbook(file2)

    sheetnames1 = wb1.sheetnames
    sheetnames2 = wb2.sheetnames

    if set(sheetnames1) != set(sheetnames2):
        messagebox.showwarning("Warning", "两个Excel文件的工作表名称不同。")
        return

    differences = []
    ABC = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O']
    for sheetname in sheetnames1:
        ws1 = wb1[sheetname]
        ws2 = wb2[sheetname]
        max_row = ws1.max_row if ws1.max_row > ws2.max_row else ws2.max_row
        max_column = ws1.max_column if ws1.max_column > ws2.max_column else ws2.max_column
        for row in range(1, max_row + 1):
            for col in range(1, max_column + 1):
                cell1 = ws1.cell(row=row, column=col)
                cell2 = ws2.cell(row=row, column=col)
                col_cell = ABC[col - 1]
                if cell1.value != cell2.value:
                    # differences.append(f"工作表：{sheetname}，行：{row}，列：{col}，文件1中的值：{cell1.value}，文件2中的值：{cell2.value}")
                    differences.append(
                        f"工作表：{sheetname}，第{row}行，第{col_cell}列，新表格中的值为：{cell1.value}，旧表格中的值为：{cell2.value}")

    if len(differences) > 0:
        current_time = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file_path = os.path.join(os.getcwd(), f'{current_time}_result.log')
        # output_file_path = os.path.join(os.getcwd(), f'output.log')
        save_to_txt_file(output_file_path, differences)
        show_compare_results(differences, output_file_path)
    else:
        messagebox.showinfo("No Differences", "两个Excel文件的内容相同。")


def show_compare_results(differences, output_file_path):
    top = Toplevel()
    top.title("比较结果")
    top.geometry("600x350")  # 设置窗口大小为 600x350

    text_widget = tk.Text(top, wrap='word')
    text_widget.pack(padx=10, pady=10)

    for difference in differences:
        text_widget.insert(tk.END, difference + '\n')
    messagebox.showinfo("比对结果", f"比较完成，结果已保存到 {output_file_path}")


def open_files():
    messagebox.showinfo("tip!", f"选择比对的新表格excel")
    file1 = filedialog.askopenfilename(title="选择第一个Excel文件")
    messagebox.showinfo("tip!", f"选择比对的旧表格excel")
    file2 = filedialog.askopenfilename(title="选择第二个Excel文件")

    if not file1 or not file2:
        messagebox.showwarning("Warning", "请选择两个Excel文件。")
        return
    compare_excel_files(file1, file2)


def main():
    root = Tk()
    root.title("Excel Compare Tool")
    # 设置窗口大小并使其居屏幕中心
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()
    window_width = 700
    window_height = 500
    x_coordinate = (screen_width - window_width) // 2
    y_coordinate = (screen_height - window_height) // 2
    root.geometry(f"{window_width}x{window_height}+{x_coordinate}+{y_coordinate}")

    label = Label(root, text="请选择两个Excel文件进行对比：")
    label.pack(pady=20)

    button = Button(root, text="选择文件", command=open_files)
    button.pack(pady=10)

    root.mainloop()


if __name__ == "__main__":
    main()
